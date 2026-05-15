from fastapi import APIRouter, HTTPException, Request, Query
from fastapi.responses import Response as FastAPIResponse
from datetime import datetime, timezone, timedelta
from typing import Optional
from io import BytesIO

from database import db
from auth import get_current_user

router = APIRouter(prefix="/api/reports", tags=["reports"])


@router.get("/sales-summary")
async def sales_summary(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    branch_id: Optional[str] = None,
    vendor_id: Optional[str] = None
):
    user = await get_current_user(request)
    bid = user["business_id"]
    query = {"business_id": bid}
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"
    if branch_id:
        query["branch_id"] = branch_id
    if vendor_id:
        query["vendedor_id"] = vendor_id

    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    total_ventas = sum(s.get("total", 0) for s in sales)
    total_iva = sum(s.get("total_iva", 0) for s in sales)
    total_descuentos = sum(s.get("descuento_global", 0) for s in sales)
    num_ventas = len(sales)

    by_payment = {}
    for s in sales:
        for p in s.get("pagos", []):
            m = p.get("metodo", "otro")
            by_payment[m] = by_payment.get(m, 0) + p.get("monto", 0)

    by_day = {}
    for s in sales:
        day = s.get("created_at", "")[:10]
        if day:
            by_day.setdefault(day, {"ventas": 0, "total": 0})
            by_day[day]["ventas"] += 1
            by_day[day]["total"] += s.get("total", 0)
    daily_chart = [{"fecha": k, "ventas": v["ventas"], "total": round(v["total"], 2)} for k, v in sorted(by_day.items())]

    return {
        "total_ventas": round(total_ventas, 2),
        "total_iva": round(total_iva, 2),
        "total_descuentos": round(total_descuentos, 2),
        "num_ventas": num_ventas,
        "promedio_venta": round(total_ventas / max(num_ventas, 1), 2),
        "por_metodo_pago": {k: round(v, 2) for k, v in by_payment.items()},
        "chart_diario": daily_chart,
    }


@router.get("/sales-by-category")
async def sales_by_category(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
):
    user = await get_current_user(request)
    query = {"business_id": user["business_id"]}
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    sales = await db.sales.find(query, {"_id": 0, "items": 1}).to_list(10000)

    # Pre-fetch all product categories to avoid N+1
    all_prod_ids = set()
    for sale in sales:
        for item in sale.get("items", []):
            pid = item.get("producto_id")
            if pid:
                all_prod_ids.add(pid)

    prod_cats = {}
    if all_prod_ids:
        prods = await db.products.find(
            {"id": {"$in": list(all_prod_ids)}}, {"_id": 0, "id": 1, "categoria_nombre": 1}
        ).to_list(len(all_prod_ids))
        prod_cats = {p["id"]: p.get("categoria_nombre", "Sin categoría") for p in prods}

    cat_totals = {}
    for sale in sales:
        for item in sale.get("items", []):
            cat = prod_cats.get(item.get("producto_id"), "Sin categoría")
            cat_totals.setdefault(cat, {"cantidad": 0, "total": 0})
            cat_totals[cat]["cantidad"] += item.get("cantidad", 0)
            cat_totals[cat]["total"] += item.get("total", 0)

    result = [{"categoria": k, "cantidad": v["cantidad"], "total": round(v["total"], 2)} for k, v in sorted(cat_totals.items(), key=lambda x: -x[1]["total"])]
    return {"categories": result}


@router.get("/sales-by-product")
async def sales_by_product(
    request: Request,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    limit: int = 20
):
    user = await get_current_user(request)
    query = {"business_id": user["business_id"]}
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    sales = await db.sales.find(query, {"_id": 0, "items": 1}).to_list(10000)
    prod_totals = {}
    for sale in sales:
        for item in sale.get("items", []):
            pid = item.get("producto_id", "")
            prod_totals.setdefault(pid, {"nombre": item.get("nombre", ""), "cantidad": 0, "total": 0})
            prod_totals[pid]["cantidad"] += item.get("cantidad", 0)
            prod_totals[pid]["total"] += item.get("total", 0)

    result = sorted(prod_totals.values(), key=lambda x: -x["total"])[:limit]
    for r in result:
        r["total"] = round(r["total"], 2)
    return {"products": result}


@router.get("/sales-by-vendor")
async def sales_by_vendor(request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None):
    user = await get_current_user(request)
    query = {"business_id": user["business_id"]}
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    sales = await db.sales.find(query, {"_id": 0, "vendedor_nombre": 1, "total": 1}).to_list(10000)
    vendor_totals = {}
    for s in sales:
        v = s.get("vendedor_nombre", "Desconocido")
        vendor_totals.setdefault(v, {"ventas": 0, "total": 0})
        vendor_totals[v]["ventas"] += 1
        vendor_totals[v]["total"] += s.get("total", 0)

    result = [{"vendedor": k, "ventas": v["ventas"], "total": round(v["total"], 2)} for k, v in sorted(vendor_totals.items(), key=lambda x: -x[1]["total"])]
    return {"vendors": result}


@router.get("/inventory-valuation")
async def inventory_valuation(request: Request):
    user = await get_current_user(request)
    products = await db.products.find(
        {"business_id": user["business_id"], "is_active": True}, {"_id": 0}
    ).to_list(10000)

    total_costo = sum(p.get("stock_actual", 0) * p.get("precio_costo", 0) for p in products)
    total_venta = sum(p.get("stock_actual", 0) * p.get("precio_venta", 0) for p in products)
    total_items = sum(p.get("stock_actual", 0) for p in products)

    by_category = {}
    for p in products:
        cat = p.get("categoria_nombre", "Sin categoría")
        by_category.setdefault(cat, {"items": 0, "costo": 0, "venta": 0})
        by_category[cat]["items"] += p.get("stock_actual", 0)
        by_category[cat]["costo"] += p.get("stock_actual", 0) * p.get("precio_costo", 0)
        by_category[cat]["venta"] += p.get("stock_actual", 0) * p.get("precio_venta", 0)

    cats = [{"categoria": k, "items": v["items"], "costo": round(v["costo"], 2), "venta": round(v["venta"], 2)} for k, v in sorted(by_category.items(), key=lambda x: -x[1]["venta"])]

    return {
        "total_productos": len(products),
        "total_items": total_items,
        "valor_costo": round(total_costo, 2),
        "valor_venta": round(total_venta, 2),
        "margen_bruto": round(total_venta - total_costo, 2),
        "por_categoria": cats,
    }


@router.get("/cash-register-history")
async def cash_register_history(request: Request, page: int = 1, limit: int = 20):
    user = await get_current_user(request)
    query = {"business_id": user["business_id"], "estado": "cerrada"}
    total = await db.cash_registers.count_documents(query)
    registers = await db.cash_registers.find(query, {"_id": 0}).sort("closed_at", -1).skip((page - 1) * limit).limit(limit).to_list(limit)
    return {"registers": registers, "total": total, "page": page}


@router.get("/invoices-summary")
async def invoices_summary(request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None):
    user = await get_current_user(request)
    query = {"business_id": user["business_id"]}
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    invoices = await db.invoices.find(query, {"_id": 0}).to_list(10000)
    by_status = {}
    by_type = {}
    for inv in invoices:
        st = inv.get("estado", "desconocido")
        by_status[st] = by_status.get(st, 0) + 1
        tp = inv.get("tipo_documento_nombre", "Otro")
        by_type.setdefault(tp, {"count": 0, "total": 0})
        by_type[tp]["count"] += 1
        by_type[tp]["total"] += inv.get("total", 0)

    return {
        "total_comprobantes": len(invoices),
        "por_estado": by_status,
        "por_tipo": [{**{"tipo": k}, **v, "total": round(v["total"], 2)} for k, v in by_type.items()],
    }


@router.get("/export/sales-excel")
async def export_sales_excel(request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None):
    user = await get_current_user(request)
    query = {"business_id": user["business_id"]}
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(["Fecha", "Vendedor", "Cliente", "Subtotal", "IVA", "Descuento", "Total", "Método Pago"])
    for s in sales:
        methods = ", ".join([p.get("metodo", "") for p in s.get("pagos", [])])
        ws.append([
            s.get("created_at", "")[:19],
            s.get("vendedor_nombre", ""),
            s.get("cliente", {}).get("nombre", "Consumidor Final"),
            s.get("subtotal_sin_iva", 0),
            s.get("total_iva", 0),
            s.get("descuento_global", 0),
            s.get("total", 0),
            methods
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FastAPIResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ventas.xlsx"}
    )


@router.get("/export/inventory-excel")
async def export_inventory_excel(request: Request):
    user = await get_current_user(request)
    products = await db.products.find(
        {"business_id": user["business_id"], "is_active": True}, {"_id": 0}
    ).to_list(10000)

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append(["Código", "Nombre", "Categoría", "Unidad", "P.Costo", "P.Venta", "IVA%", "Stock", "Stk.Mín", "Ubicación"])
    for p in products:
        ws.append([
            p.get("codigo_interno", ""),
            p.get("nombre", ""),
            p.get("categoria_nombre", ""),
            p.get("unidad_medida", ""),
            p.get("precio_costo", 0),
            p.get("precio_venta", 0),
            p.get("iva_porcentaje", 0),
            p.get("stock_actual", 0),
            p.get("stock_minimo", 0),
            p.get("ubicacion", ""),
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FastAPIResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=inventario.xlsx"}
    )


@router.get("/export/sales-pdf")
async def export_sales_pdf(request: Request, date_from: Optional[str] = None, date_to: Optional[str] = None):
    user = await get_current_user(request)
    query = {"business_id": user["business_id"]}
    if date_from:
        query.setdefault("created_at", {})["$gte"] = date_from
    if date_to:
        query.setdefault("created_at", {})["$lte"] = date_to + "T23:59:59"

    sales = await db.sales.find(query, {"_id": 0}).sort("created_at", -1).to_list(10000)
    business = await db.businesses.find_one({"id": user["business_id"]}, {"_id": 0})

    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(0, 10, business.get("nombre_comercial", "Reporte") if business else "Reporte", ln=True, align="C")
    pdf.set_font("Helvetica", "", 10)
    period = ""
    if date_from:
        period += f"Desde: {date_from} "
    if date_to:
        period += f"Hasta: {date_to}"
    if period:
        pdf.cell(0, 6, period, ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Helvetica", "B", 8)
    pdf.cell(35, 7, "Fecha", border=1)
    pdf.cell(35, 7, "Vendedor", border=1)
    pdf.cell(45, 7, "Cliente", border=1)
    pdf.cell(25, 7, "Subtotal", border=1, align="C")
    pdf.cell(20, 7, "IVA", border=1, align="C")
    pdf.cell(25, 7, "Total", border=1, align="C")
    pdf.ln()

    pdf.set_font("Helvetica", "", 7)
    total_sum = 0
    for s in sales[:200]:
        pdf.cell(35, 6, s.get("created_at", "")[:16], border=1)
        pdf.cell(35, 6, str(s.get("vendedor_nombre", ""))[:20], border=1)
        pdf.cell(45, 6, str(s.get("cliente", {}).get("nombre", "CF"))[:25], border=1)
        pdf.cell(25, 6, f"${s.get('subtotal_sin_iva', 0):.2f}", border=1, align="C")
        pdf.cell(20, 6, f"${s.get('total_iva', 0):.2f}", border=1, align="C")
        pdf.cell(25, 6, f"${s.get('total', 0):.2f}", border=1, align="C")
        pdf.ln()
        total_sum += s.get("total", 0)

    pdf.set_font("Helvetica", "B", 10)
    pdf.ln(3)
    pdf.cell(0, 8, f"Total General: ${total_sum:.2f} ({len(sales)} ventas)", align="R")

    return FastAPIResponse(
        content=bytes(pdf.output()),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_ventas.pdf"}
    )
