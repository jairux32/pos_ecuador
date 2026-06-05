from fastapi import APIRouter, HTTPException, Request, UploadFile, File, Query
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone
from bson import ObjectId
import uuid

from database import db
from auth import get_current_user
from utils.ecuador import UNIDADES_MEDIDA, TASAS_IVA, MOTIVOS_AJUSTE_INVENTARIO
from utils.storage import upload_file
from routes.audit_routes import log_audit

router = APIRouter(prefix="/api/inventory", tags=["inventory"])


class ProductCreate(BaseModel):
    codigo_interno: Optional[str] = None
    codigo_barras: Optional[str] = None
    nombre: str
    descripcion: Optional[str] = ""
    categoria_id: Optional[str] = None
    categoria_nombre: Optional[str] = "General"
    unidad_medida: str = "Unidad"
    precio_costo: float = 0
    precio_venta: float = 0
    iva_porcentaje: float = 15
    stock_actual: float = 0
    stock_minimo: float = 0
    stock_maximo: float = 1000
    ubicacion: Optional[str] = ""
    proveedor_id: Optional[str] = None
    imagen_path: Optional[str] = None


class ProductUpdate(BaseModel):
    codigo_interno: Optional[str] = None
    codigo_barras: Optional[str] = None
    nombre: Optional[str] = None
    descripcion: Optional[str] = None
    categoria_id: Optional[str] = None
    categoria_nombre: Optional[str] = None
    unidad_medida: Optional[str] = None
    precio_costo: Optional[float] = None
    precio_venta: Optional[float] = None
    iva_porcentaje: Optional[float] = None
    stock_actual: Optional[float] = None
    stock_minimo: Optional[float] = None
    stock_maximo: Optional[float] = None
    ubicacion: Optional[str] = None
    proveedor_id: Optional[str] = None
    imagen_path: Optional[str] = None
    is_active: Optional[bool] = None


class StockAdjustment(BaseModel):
    producto_id: str
    cantidad: float
    tipo: str  # entrada, salida, ajuste
    motivo: str
    notas: Optional[str] = ""


class CategoryCreate(BaseModel):
    nombre: str


@router.get("/units")
async def get_units():
    return {"units": UNIDADES_MEDIDA}


@router.get("/iva-rates")
async def get_iva_rates():
    return {"rates": TASAS_IVA}


@router.get("/adjustment-reasons")
async def get_adjustment_reasons():
    return {"reasons": MOTIVOS_AJUSTE_INVENTARIO}


@router.get("/categories")
async def get_categories(request: Request):
    user = await get_current_user(request)
    categories = await db.categories.find(
        {"business_id": user["business_id"]}, {"_id": 0}
    ).to_list(200)
    return {"categories": categories}


@router.post("/categories")
async def create_category(body: CategoryCreate, request: Request):
    user = await get_current_user(request)
    cat_doc = {
        "id": str(uuid.uuid4()),
        "business_id": user["business_id"],
        "nombre": body.nombre,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.categories.insert_one(cat_doc)
    return {"id": cat_doc["id"], "nombre": cat_doc["nombre"]}


@router.get("/products")
async def get_products(
    request: Request,
    search: Optional[str] = None,
    category: Optional[str] = None,
    low_stock: Optional[bool] = False,
    page: int = 1,
    limit: int = 50
):
    user = await get_current_user(request)
    query = {"business_id": user["business_id"], "is_active": True}

    if search:
        query["$or"] = [
            {"nombre": {"$regex": search, "$options": "i"}},
            {"codigo_interno": {"$regex": search, "$options": "i"}},
            {"codigo_barras": {"$regex": search, "$options": "i"}},
        ]
    if category:
        query["categoria_nombre"] = category

    if low_stock:
        query["$expr"] = {"$lte": ["$stock_actual", "$stock_minimo"]}

    skip = (page - 1) * limit
    total = await db.products.count_documents(query)
    products = await db.products.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)

    return {
        "products": products,
        "total": total,
        "page": page,
        "pages": (total + limit - 1) // limit
    }


@router.post("/products")
async def create_product(body: ProductCreate, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["superadmin", "administrador", "bodeguero"]:
        raise HTTPException(status_code=403, detail="No tiene permisos para crear productos")

    product_id = str(uuid.uuid4())
    codigo_interno = body.codigo_interno or f"P-{product_id[:8].upper()}"

    product_doc = {
        "id": product_id,
        "business_id": user["business_id"],
        "codigo_interno": codigo_interno,
        "codigo_barras": body.codigo_barras or "",
        "nombre": body.nombre,
        "descripcion": body.descripcion or "",
        "categoria_id": body.categoria_id,
        "categoria_nombre": body.categoria_nombre or "General",
        "unidad_medida": body.unidad_medida,
        "precio_costo": body.precio_costo,
        "precio_venta": body.precio_venta,
        "iva_porcentaje": body.iva_porcentaje,
        "stock_actual": body.stock_actual,
        "stock_minimo": body.stock_minimo,
        "stock_maximo": body.stock_maximo,
        "ubicacion": body.ubicacion or "",
        "proveedor_id": body.proveedor_id,
        "imagen_path": body.imagen_path,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    await db.products.insert_one(product_doc)

    await db.inventory_movements.insert_one({
        "id": str(uuid.uuid4()),
        "business_id": user["business_id"],
        "producto_id": product_id,
        "producto_nombre": body.nombre,
        "tipo": "entrada",
        "cantidad": body.stock_actual,
        "stock_anterior": 0,
        "stock_nuevo": body.stock_actual,
        "motivo": "Stock inicial",
        "usuario_id": user["id"],
        "usuario_nombre": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    })

    product_doc.pop("_id", None)
    await log_audit(user["business_id"], user["id"], user.get("name",""), "crear_producto", "producto", product_id, f"Producto: {body.nombre}", request.client.host if request.client else "")
    return product_doc


@router.get("/products/{product_id}")
async def get_product(product_id: str, request: Request):
    user = await get_current_user(request)
    product = await db.products.find_one(
        {"id": product_id, "business_id": user["business_id"]}, {"_id": 0}
    )
    if not product or not product.get("is_active", True):
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    return product


@router.put("/products/{product_id}")
async def update_product(product_id: str, body: ProductUpdate, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["superadmin", "administrador", "bodeguero"]:
        raise HTTPException(status_code=403, detail="No tiene permisos para editar productos")

    product = await db.products.find_one(
        {"id": product_id, "business_id": user["business_id"]}
    )
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    update_data = {k: v for k, v in body.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    await db.products.update_one(
        {"id": product_id}, {"$set": update_data}
    )
    updated = await db.products.find_one({"id": product_id}, {"_id": 0})
    await log_audit(
        user["business_id"], user["id"], user.get("name", ""),
        "editar_producto", "producto", product_id,
        f"Producto: {updated.get('nombre', '')}",
        request.client.host if request.client else "",
    )
    return updated


@router.delete("/products/{product_id}")
async def delete_product(product_id: str, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["superadmin", "administrador"]:
        raise HTTPException(status_code=403, detail="No tiene permisos para eliminar productos")

    product = await db.products.find_one(
        {"id": product_id, "business_id": user["business_id"]}
    )
    result = await db.products.update_one(
        {"id": product_id, "business_id": user["business_id"]},
        {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    await log_audit(
        user["business_id"], user["id"], user.get("name", ""),
        "eliminar_producto", "producto", product_id,
        f"Producto: {product.get('nombre', '') if product else product_id}",
        request.client.host if request.client else "",
    )
    return {"message": "Producto eliminado"}


@router.post("/stock-adjustment")
async def adjust_stock(body: StockAdjustment, request: Request):
    user = await get_current_user(request)
    if user["role"] not in ["superadmin", "administrador", "bodeguero"]:
        raise HTTPException(status_code=403, detail="No tiene permisos para ajustar stock")

    product = await db.products.find_one(
        {"id": body.producto_id, "business_id": user["business_id"]}
    )
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    stock_anterior = product["stock_actual"]
    if body.tipo == "entrada":
        nuevo_stock = stock_anterior + body.cantidad
    elif body.tipo == "salida":
        nuevo_stock = stock_anterior - body.cantidad
        if nuevo_stock < 0:
            raise HTTPException(status_code=400, detail="Stock insuficiente")
    else:
        nuevo_stock = body.cantidad

    await db.products.update_one(
        {"id": body.producto_id},
        {"$set": {"stock_actual": nuevo_stock, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )

    movement_doc = {
        "id": str(uuid.uuid4()),
        "business_id": user["business_id"],
        "producto_id": body.producto_id,
        "producto_nombre": product["nombre"],
        "tipo": body.tipo,
        "cantidad": body.cantidad,
        "stock_anterior": stock_anterior,
        "stock_nuevo": nuevo_stock,
        "motivo": body.motivo,
        "notas": body.notas or "",
        "usuario_id": user["id"],
        "usuario_nombre": user.get("name", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.inventory_movements.insert_one(movement_doc)
    movement_doc.pop("_id", None)
    return movement_doc


@router.get("/movements")
async def get_movements(
    request: Request,
    product_id: Optional[str] = None,
    page: int = 1,
    limit: int = 50
):
    user = await get_current_user(request)
    query = {"business_id": user["business_id"]}
    if product_id:
        query["producto_id"] = product_id

    skip = (page - 1) * limit
    total = await db.inventory_movements.count_documents(query)
    movements = await db.inventory_movements.find(
        query, {"_id": 0}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)

    return {"movements": movements, "total": total, "page": page}


@router.get("/low-stock")
async def get_low_stock(request: Request):
    user = await get_current_user(request)
    products = await db.products.find(
        {
            "business_id": user["business_id"],
            "is_active": True,
            "$expr": {"$lte": ["$stock_actual", "$stock_minimo"]}
        },
        {"_id": 0}
    ).to_list(200)
    return {"products": products, "count": len(products)}


@router.post("/upload-image")
async def upload_product_image(request: Request, file: UploadFile = File(...)):
    await get_current_user(request)
    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Solo se permiten imágenes")
    data = await file.read()
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Máximo 5MB")
    try:
        path = upload_file(data, file.filename, file.content_type, "products")
        return {"path": path}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")


@router.post("/import-csv")
async def import_products_csv(request: Request, file: UploadFile = File(...)):
    user = await get_current_user(request)
    if user["role"] not in ["superadmin", "administrador", "bodeguero"]:
        raise HTTPException(status_code=403, detail="No tiene permisos para importar")

    content = await file.read()
    filename = file.filename or "import"

    imported = 0
    errors = []

    try:
        if filename.endswith(".xlsx") or filename.endswith(".xls"):
            import openpyxl
            from io import BytesIO
            wb = openpyxl.load_workbook(BytesIO(content))
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            headers = [str(c.value or "").strip().lower() for c in ws[1]]
        else:
            import csv
            from io import StringIO
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            headers = [h.strip().lower() for h in (reader.fieldnames or [])]
            rows = None
            raw_rows = list(reader)

        field_map = {
            "nombre": ["nombre", "name", "producto", "descripcion"],
            "codigo_interno": ["codigo_interno", "codigo", "code", "sku"],
            "codigo_barras": ["codigo_barras", "barcode", "ean"],
            "precio_costo": ["precio_costo", "costo", "cost", "precio_compra"],
            "precio_venta": ["precio_venta", "precio", "price", "pvp"],
            "stock_actual": ["stock_actual", "stock", "cantidad", "qty"],
            "stock_minimo": ["stock_minimo", "minimo", "min_stock"],
            "categoria_nombre": ["categoria", "category", "categoria_nombre"],
            "unidad_medida": ["unidad_medida", "unidad", "unit"],
            "iva_porcentaje": ["iva_porcentaje", "iva", "tax"],
        }

        def find_col(field):
            for alias in field_map.get(field, []):
                if alias in headers:
                    return headers.index(alias) if rows is not None else alias
            return None

        col_indices = {f: find_col(f) for f in field_map}
        nombre_col = col_indices.get("nombre")
        if nombre_col is None:
            raise HTTPException(status_code=400, detail="No se encontró columna 'nombre' en el archivo. Columnas: " + ", ".join(headers))

        data_rows = rows if rows is not None else raw_rows

        for i, row in enumerate(data_rows):
            try:
                if rows is not None:
                    get_val = lambda col: str(row[col]).strip() if col is not None and col < len(row) and row[col] is not None else ""
                else:
                    get_val = lambda col: str(row.get(col, "")).strip() if col else ""

                nombre = get_val(col_indices["nombre"])
                if not nombre:
                    continue

                def safe_float(col, default=0):
                    v = get_val(col)
                    try:
                        return float(v) if v else default
                    except (ValueError, TypeError):
                        return default

                codigo_interno = get_val(col_indices["codigo_interno"]) or f"IMP-{str(uuid.uuid4())[:8].upper()}"

                existing = await db.products.find_one({
                    "business_id": user["business_id"],
                    "codigo_interno": codigo_interno,
                    "is_active": True,
                })
                if existing:
                    errors.append(
                        f"Fila {i + 2}: codigo_interno '{codigo_interno}' ya existe"
                    )
                    continue

                product_doc = {
                    "id": str(uuid.uuid4()),
                    "business_id": user["business_id"],
                    "codigo_interno": codigo_interno,
                    "codigo_barras": get_val(col_indices["codigo_barras"]),
                    "nombre": nombre,
                    "descripcion": "",
                    "categoria_nombre": get_val(col_indices["categoria_nombre"]) or "General",
                    "unidad_medida": get_val(col_indices["unidad_medida"]) or "Unidad",
                    "precio_costo": safe_float(col_indices["precio_costo"]),
                    "precio_venta": safe_float(col_indices["precio_venta"]),
                    "iva_porcentaje": safe_float(col_indices["iva_porcentaje"], 15),
                    "stock_actual": safe_float(col_indices["stock_actual"]),
                    "stock_minimo": safe_float(col_indices["stock_minimo"]),
                    "stock_maximo": 1000,
                    "ubicacion": "",
                    "proveedor_id": None,
                    "imagen_path": None,
                    "is_active": True,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
                await db.products.insert_one(product_doc)
                imported += 1
            except Exception as e:
                errors.append(f"Fila {i + 2}: {str(e)}")

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error procesando archivo: {str(e)}")

    return {
        "imported": imported,
        "errors": errors[:20],
        "total_errors": len(errors),
        "message": f"Se importaron {imported} productos" + (f" con {len(errors)} errores" if errors else "")
    }


@router.get("/export-template")
async def export_template():
    from openpyxl import Workbook
    from io import BytesIO
    from fastapi.responses import Response as FastAPIResponse

    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Productos"
    ws.append(["nombre", "codigo_interno", "codigo_barras", "categoria", "unidad_medida", "precio_costo", "precio_venta", "iva_porcentaje", "stock_actual", "stock_minimo"])
    ws.append(["Arroz 1kg", "ARR-001", "7861234567890", "Alimentos", "Unidad", 0.90, 1.25, 0, 100, 10])
    ws.append(["Aceite 1L", "ACE-001", "7861234567891", "Alimentos", "Unidad", 2.50, 3.25, 15, 50, 5])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return FastAPIResponse(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_productos.xlsx"}
    )
